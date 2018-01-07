#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import codecs       # utf8 codecs
import csv          # read/write csv files
import os
import re           # regular expression
import shutil       # copy files
import sqlite3      # lightweight database
import subprocess   # execute shell commands
import sys          # system
import traceback    # dealing with exception
import xlsxwriter   # export xlsx
import logging      # for debug
#from PyQt5.Qt import QObject
from openpyxl import Workbook, worksheet, load_workbook
from platform import uname

# format the typesetting of names
class Genlist(object):

    def __init__(self, parent=None):
        super(Genlist, self).__init__()

    def resource_path(self, relative):
        """
        resource_path(relative)
        =======================
        For pyinstaller

        ref:  http://stackoverflow.com/questions/7674790/bundling-data-files-with-pyinstaller-onefile
        """
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative)
        else:
            return(relative)
        return os.path.join(os.path.abspath("."), relative)

    def fmtnameNew(self, fullname, format_type='markdown', italic_b="*", italic_e="*", withSpAuthor = True, doformat=True, split=True):

        ### clean fullname: remove multiple spaces
        #fullname = re.sub('  ',' ', fullname)

        ### format
        if doformat == False:
            italic_b = ''
            italic_e = ''
        else:
            if fullname is None or fullname is '':
                print('Usage: fmtname(fullname_with_author)')
                return
            if format_type == 'markdown':
                italic_b = '*'; italic_e = '*'
            elif format_type == 'html':
                italic_b = '<i>'; italic_e = '</i>'
            elif format_type == 'custom':
                italic_b; italic_e
            else:
                print('Unsupported format type: %s (only support markdown and html)' % format_type)
                return

        ### processing subordinate
        fullname = re.sub('  ',' ', fullname)
        fullname_split = fullname.split(' ')
        length_fullname = len(fullname_split)


        epithetLst = []
        epithet = ''
        author_start = ''
        cross_type = ''
        subordinate_status = []

        subrank = ['subsp.', 'ssp.', 'var.', 'fo.', 'cv.', '×', 'x']
        autonym = 'false'
        for s in range(0,len(subrank)):
            if subrank[s] in fullname_split:
                subrank_idx = fullname_split.index(subrank[s])
                if subrank[s] == '×' or subrank[s] == 'x':
                    # 把 x 取代成 ×
                    subrank[s] = '×'
                    fullname_split[subrank_idx] = '×'
                    # × 後面有跟著屬名的
                    # ex: Genus epithet × Genus epithet2
                    if re.search('[A-Z].*', fullname_split[subrank_idx+1]):
                        subordinate_status.append([subrank_idx, subrank[s]])
                        cross_type = 1
                    else:
                        # 沒有屬名的 case
                        fullname_split[subrank_idx+1] = '×' + fullname_split[subrank_idx+1]
                        fullname_split.remove(subrank[s])
                else:
                    subordinate_status.append([subrank_idx, subrank[s]])

                # check for autonyms
                epithet = fullname_split[1]
                sub_epithet = fullname_split[subrank_idx + 1]
                if epithet == sub_epithet:
                    autonym = True

        fname_sp = italic_b + ' '.join(str(item) for item in fullname_split[0:2])+ italic_e

        #### 有種下階層的
        subordinate_status = sorted(subordinate_status)

        if cross_type == 1:
            sidx = int(subordinate_status[0][0])
            subepithet = italic_b + fullname_split[sidx+1] + ' ' + fullname_split[sidx+2] + italic_e
            speciesAuthors = ' '.join(str(item) for item in fullname_split[2:sidx])
            subepithetAuthors = ' '.join(str(item) for item in fullname_split[sidx+3:len(fullname_split)])
            fullnameNoAuthors = ' '.join([fname_sp, '×',subepithet])
            fullnameWithAuthors = ' '.join([fname_sp,speciesAuthors,'×',subepithet,subepithetAuthors])

        # autonym
        elif autonym == True:
            subrank_name = fullname_split[subrank_idx]
            subepithet = italic_b + fullname_split[subordinate_status[0][0]+1] + italic_e
            fullnameWithAuthors = ' '.join([fname_sp, ' '.join(fullname_split[2:subrank_idx]), subrank_name, \
                  subepithet])
            fullnameNoAuthors = ' '.join([fname_sp, subrank_name, subepithet])

        # 只有一個種下階層的
        elif len(subordinate_status) == 1:
            subrank_name = fullname_split[subrank_idx]
            subepithet = italic_b + fullname_split[subordinate_status[0][0]+1] + italic_e

            if withSpAuthor == True:
                fullnameWithAuthors = ' '.join([fname_sp, ' '.join(fullname_split[2:subrank_idx]), subrank_name, \
                  subepithet, ' '.join(fullname_split[subrank_idx+2:len(fullname_split)])])
            elif withSpAuthor == False:
                fullnameWithAuthors = ' '.join([fname_sp, subrank_name, \
                  subepithet, ' '.join(fullname_split[subrank_idx+2:len(fullname_split)])])
            fullnameNoAuthors = ' '.join([fname_sp, subrank_name, subepithet])
        else:
            authors = fullname_split[2:length_fullname]
            authors_join = ' '.join(authors)
            fullnameNoAuthors = ' '.join([fname_sp])
            fullnameWithAuthors = ' '.join([fullnameNoAuthors, authors_join])
        # returns
        if split == False and withSpAuthor == False:
            return(fullnameNoAuthors)
        elif split == False:
            return(fullnameWithAuthors)
        elif split == True:
            return([fullnameNoAuthors, fullnameWithAuthors])
        else:
            print("Invalid split option. Please choose 'True' or 'False'")

    def fmtname(self, fullname, italic_b="*", italic_e="*", format_type='markdown', doformat=True, split=True):
        """
        fmtname(fullname, italic_b="*", italic_e="*", format_type='markdown', doformat=True, split=True)
        =======
        This function will format a scientific name into following rules:
        1. genus and species name are italicized
        2. subrank (ex: subsp., var., etc.) is not italicized

        html tag example <i></i> means italic fonts):
        <i>Castanopsis cuspidata</i> (Thunb.) Schottky var. <i>carlesii</i> (Hemsl.) T.Yamaz.'

        fullname:
        --------
        fullname, ex: 'Castanopsis cuspidata (Thunb.) Schottky var. carlesii (Hemsl.) T.Yamaz.'

        format_type:
        ------------
        markdown, html or custom, default is markdown

        italic_b:
        ---------
        beginning tag of italic, default is "*"

        italic_e:
        ---------
        end tag of italic, default is "*"

        split:
        --------
        After the fullname is formatted, fmtname will return a list:
        [formatted_fullname, authors]

        If split option is true, fmtname will return a string of fullname and authors.
        """
        if fullname is None or fullname is '':
            print('Usage: fmtname(fullname_with_author)')
            return
        if format_type == 'markdown':
            italic_b = '*'; italic_e = '*'
        elif format_type == 'html':
            italic_b = '<i>'; italic_e = '</i>'
        elif format_type == 'custom':
            italic_b; italic_e
        else:
            print('Unsupported format type: %s (only support markdown and html)' % format_type)
            return
        if doformat == False:
            italic_b = ''
            italic_e = ''
        # remove trailing spaces
        fullname_split = fullname.split(' ')
        while '' in fullname_split:
            fullname_split.remove('')
        length_fullname = len(fullname_split)
        subordinate_status = []

        if '×' in fullname_split:
            cross_idx = fullname_split.index('×')
            subordinate_status.append([cross_idx, '×'])
        if 'subsp.' in fullname_split:
            subsp_idx = fullname_split.index('subsp.')
            subordinate_status.append([subsp_idx, 'subsp.'])
        if 'ssp.' in fullname_split:
            subsp_idx = fullname_split.index('ssp.')
            subordinate_status.append([subsp_idx, 'ssp.'])
        if 'var.' in fullname_split:
            var_idx = fullname_split.index('var.')
            subordinate_status.append([var_idx, 'var.'])
        if 'fo.' in fullname_split:
            fo_idx = fullname_split.index('fo.')
            subordinate_status.append([fo_idx, 'fo.'])
        if 'cv.' in fullname_split:
            cv_idx = fullname_split.index('cv.')
            subordinate_status.append([cv_idx, 'cv.'])
        subordinate_status = sorted(subordinate_status)
        fname_sp = italic_b + ' '.join(str(item) for item in fullname_split[0:2])+ italic_e
        next_s = ''
        epithetLst = []
        epithet = ''
        authors_start = ''
        if len(subordinate_status) >= 1:
            for v in range(len(subordinate_status)):
                if subordinate_status[v][1] == '×':
                    if subordinate_status[v][0] == 0:
                        fname_sp =  italic_b + '×' + ' '.join(str(item) for item in fullname_split[1:3])+ italic_e
                        authors_start = 3
                    elif subordinate_status[v][0] == 1:
                        fname_sp = italic_b + ' '.join(str(item) for item in fullname_split[0:3]) + italic_e
                        authors_start = 3
                    elif subordinate_status[v][0] > 1:
                        next_s = fullname_split[subordinate_status[v][0]+1]
                        before_s = fullname_split[subordinate_status[v][0]-1]
                        if next_s == 'var.' or next_s == 'subsp.' or next_s == 'fo.' or before_s == 'cv.':
                            epithet = '×'# + next_s + ' ' + italic_b + \
                                #fullname_split[subordinate_status[v][0]+2] + italic_e
                            authors_start = subordinate_status[v][0] + 3
                        elif before_s == 'var.' or before_s == 'subsp.' or before_s == 'fo.' or before_s == 'cv.':
                            epithet = ''
                            authors_start = subordinate_status[v][0] + 3
                        elif re.search('^[A-Z].*', next_s):
                            epithet = '×' + italic_b + fullname_split[subordinate_status[v][0]+2] + italic_e
                            authors_start = subordinate_status[v][0] + 3
                elif len(subordinate_status) >= 1:
                    if fullname_split[subordinate_status[v][0]+1] == '×':
                        sub_epithet = '×' + fullname_split[subordinate_status[v][0]+2]
                    else:
                        sub_epithet = fullname_split[subordinate_status[v][0]+1]
                    epithet = subordinate_status[v][1] + ' ' + italic_b + sub_epithet + italic_e
                epithetLst.append(epithet)
            # authors
            if authors_start is '':
                authors_start = subordinate_status[-1][0] + 2
            authors = fullname_split[authors_start:length_fullname]
            authors_join = ' '.join(authors)
            if epithet == '':
                fnameLst = fname_sp + ' ' + ' '.join(epithetLst)
            else:
                fnameLst = fname_sp + ' ' + ' '.join(epithetLst)
        else:
            authors = fullname_split[2:length_fullname]
            authors_join = ' '.join(authors)
            fnameLst = fname_sp
        # ex should be italic?
        # NEEDS TO BE CONFIRMED
        # authors_join = re.sub(' ex ', ' ' + italic_b + 'ex' + italic_e + ' ', authors_join)
        if split == False:
            return(fnameLst + ' ' + authors_join)
        elif split == True:
            return([fnameLst, authors_join])
        else:
            print("Invalid split option. Please choose 'True' or 'False'")


    def pandocConvert(self, oformat='docx', ofile_prefix='output'):
        try:
            inpfile = ofile_prefix + '.md'
            outfile = ofile_prefix + '.' + oformat
            if uname()[0] == 'Windows':
                subprocess.Popen([self.resource_path('pandoc'), '-f', 'markdown', '-t', 'docx', inpfile, '-o', outfile], shell=True)
            elif uname()[0] == 'Darwin':
                pandocPath = os.path.dirname(sys.executable)
                subprocess.Popen([os.path.join(pandocPath, 'pandoc'), inpfile, '-o', outfile])
            else:
                subprocess.Popen([self.resource_path('pandoc'), inpfile, '-o', outfile])
        except BaseException as e:
            print(str(e))

    def dbExecuteSQL(self, schema, dbfile, show_results=False):
        """
        dbExcuteSQL(schema, dbfile, show_results=False)
        ===========
        dbfile
        """
        conn = sqlite3.connect(dbfile)
        curs = conn.cursor()
        curs.execute(schema)
        if show_results == True:
            output = curs.fetchall()
            return(output)
        else:
            conn.commit()
            print("Execute '%s' successfully" % schema)
        conn.close()

    def dbImportTable(self, table_name, csvfile, dbfile):
        conn = sqlite3.connect(dbfile)
        curs = conn.cursor()
        with open(csvfile, newline='', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='|')
            for row in reader:
                insert_db = '''
                INSERT INTO %s (
                    family,
                    family_cname,
                    cname,
                    name,
                    fullname,
                    plant_type,
                    endemic,
                    iucn_category,
                    source)
                VALUES ("%s", "%s", "%s", "%s", "%s", %s,  %s, "%s", "%s");
                ''' % (table_name, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8])
                curs.execute(insert_db)
                i=i+1
            conn.commit()
        conn.close()

    def dbGetsp(self, table_name, dbfile):
        conn = sqlite3.connect(dbfile)
        curs = conn.cursor()
        get_splist_sql = '''SELECT * FROM %s ORDER BY family,name;''' % table_name
        curs.execute(get_splist_sql)
        get_splist_result = curs.fetchall()
        conn.commit()
        return(get_splist_result)
        conn.close()

    def importTable(self, dbfile, table_name, import_file, isFile=True):
        conn = sqlite3.connect(dbfile)
        curs = conn.cursor()
        curs.execute('DROP TABLE IF EXISTS %s;' % table_name)
        conn.commit()
        sample_create = '''
        CREATE TABLE %s (
          local_name varchar
        ); ''' % table_name
        curs.execute(sample_create)
        if isFile == True:
            m_lists = []
            with codecs.open(import_file, 'r', 'utf-8') as f:
                m_lists += f.read().splitlines()
            f.close()
        else:
            m_lists = import_file
        for row in range(len(m_lists)):
            zhname = re.sub(' ', '', m_lists[row])
            zhname = re.sub('\ufeff', '', zhname)
            # substitute 台 to 臺
            zhname = re.sub(u'台([灣|北|中|西|南|東])',r'臺\1', zhname)
            # pass the empty lines
            if zhname != '':
                insert_db = '''
                INSERT INTO %s (local_name) VALUES ("%s");
                ''' % (table_name, zhname)
            curs.execute(insert_db)
        conn.commit()

    def combineChecklists(self, dbfile, checklists):
        """
        dbfile: sqlite database
        checklists: list of checklists
        """
        m_lists = []
        table_name_lists = []
        for files in range(len(checklists)):
            with codecs.open(checklists[files], 'r', 'utf-8') as f:
                # import all the local names in checklists and merge
                # them into one list
                m_lists += f.read().splitlines()
                # import checklists into different tables
                checklist_filename = os.path.split(checklists[files])[1]
                checklist_tablename = str.split(checklist_filename, '.')[0]
                table_name_lists.append(checklist_tablename)
                self.importTable(dbfile, checklist_tablename, checklists[files])
            f.close()
        # create a union table includes all the species
        self.importTable(dbfile, 'tmp_union', m_lists, isFile=False)
        # update
        for t in range(len(table_name_lists)):
            add_column_sql = '''ALTER TABLE tmp_union ADD COLUMN %s varchar;''' % table_name_lists[t]
            self.dbExecuteSQL(add_column_sql, dbfile)
            update_sql = '''
            update tmp_union set %s = '+' WHERE EXISTS (SELECT * FROM %s WHERE %s.local_name = tmp_union.local_name);
            ''' % ( table_name_lists[t], table_name_lists[t], table_name_lists[t] )
            self.dbExecuteSQL(update_sql, dbfile)
        output = self.dbExecuteSQL('''SELECT distinct * FROM tmp_union;''', dbfile, show_results = True)
        return(output)

    def listToXls(self, import_list, name_italic_col, xls_file):
        """
        listToXls: write list to excel
        ==============================

        import_list
        -----------

        name_italic_col
        ---------------

        """
        try:
            logging.basicConfig(filename='/tmp/checklister.log',level=logging.DEBUG)
            xls_subname = str.split(os.path.split(xls_file)[1], '.')[-1]
            if xls_subname != 'xlsx':
                xls_file = str(xls_file) + '.xlsx'
            xlsBase = re.sub('.xls[x|]$','',xls_file)
            logging.info(xlsBase)
            #with codecs.open(xlsBase+'.csv', 'w+', 'utf-8') as f:
            #    for row in range(len(import_list)):
            #        f.write('\t'.join(import_list[row]))
            wb = xlsxwriter.Workbook(xls_file)
            ws = wb.add_worksheet()
            # rich formats
            italic = wb.add_format({'italic': True})
            default = wb.add_format()
            for row in range(len(import_list)):
                for col in range(len(import_list[row])):
                    if col == int(name_italic_col) and row > 0:
                        input_name = self.fmtname(import_list[row][col], format_type='custom', italic_b='^', italic_e='$', split=False)
                        formattedLst = []
                        splitted_name = input_name.split('^')
                        splitted_name.remove('')
                        for i in range(len(splitted_name)):
                            if re.search('$', splitted_name[i]):
                                should_italic = splitted_name[i].split('$')
                                a = [italic, should_italic[0], default, should_italic[1]]
                                for r in range(len(a)):
                                    formattedLst.append(a[r])
                        ws.write_rich_string(row, col, *formattedLst)
                    else:
                        ws.write(row,col,import_list[row][col])
            wb.close()
        except BaseException as e:
            print(str(e))

    def fmtExcelNames(self, original, outputfile, name_col_num, header=True):
        try:
            openpyxl_wb = load_workbook(original, read_only=True)
            work_sheet = openpyxl_wb.get_sheet_names()[0]
            curr_sheet = openpyxl_wb[work_sheet]
            # TODO deprecated function or class get_highest_row (Use the max_row property)
            max_col = curr_sheet.get_highest_column()
            max_row = curr_sheet.get_highest_row()
            # header
            wsLst = [['orignal', 'formatted']]
            if name_col_num > max_col:
                print(u'The column number exceed the maximum limit')
            else:
                for i in range(1, max_row):
                    if header == True:
                        i = i+1
                    d = curr_sheet.cell(row=i, column=name_col_num).value
                    wsLst.append([d,d])
            self.listToXls(wsLst, 1, outputfile)
        except BaseException as e:
            print(str(e))

    def adjCharTai(self, zhnameWithTai):
        '''
        adjCharTai
        ----------

        處理「台」和「臺」的問題
        '''
        try:
           zhname = re.sub(u'台([灣|北|中|西|南|東])',r'臺\1', zhnameWithTai)
           return(zhname)
        except BaseException as e:
            print(str(e))

    def expCombList(self, sqlite_db, current_db_table, tobe_combined_lists, exportExcel):
        """
        Combine multiple checklists
        ===========================


        Arguments
        ---------
        tobo_combined_lists: list()
        組合不同的名錄，並標示出各名錄間出現的物種

        sqlite_db:


        """
        try:
            tobe_combined_files = ','.join(tobe_combined_lists)
            self.combineChecklists(sqlite_db, tobe_combined_lists)
            #current_db_table = self.checkDB()
            # print(current_db_table)
            if current_db_table == 'dao_bnamelist':
                db_fullname = 'name'
                iucn = 'consv_status'
                list_type = 'family'
            else:
                db_fullname = 'fullname'
                iucn = 'iucn_category'
                list_type = 'plant_type,family'
                fetch_combined_sql = '''
                SELECT 
                    distinct 
                    d.family,
                    d.family_cname,
                    d.%s,
                    d.endemic,
                    d.%s,
                    u.*
                FROM
                    tmp_union u, %s d
                WHERE
                    u.local_name = d.cname order by %s
                ''' % (db_fullname, iucn, current_db_table, list_type)
                combined_table = self.dbExecuteSQL(fetch_combined_sql, sqlite_db, show_results=True)
            header = ['family', 'family_cname', 'fullname', 'endemic', iucn, 'common name']
            for i in range(len(tobe_combined_lists)):
                fname = str.split(os.path.split(tobe_combined_lists[i])[1], '.')
                header.insert(7+i, fname[0])
            header = tuple(header)
            combined_table.insert(0, header)
            self.combined_checklists = combined_table
            self.listToXls(combined_table, 2, exportExcel)
        except BaseException as e:
            print(str(e))


    def genEngine(self, dbfile, dbtable, inputfile, oformat='docx', ofile_prefix='output'):
        """
        dbfile
        ------
        sqlite database file

        dbtable
        -------

        inputfile
        ---------
        oformat
        ofile_prefix
        """
        # check for input parameters
        #if os.path(dbfile) is True:
        #    pass
        #else:
        #    print(dbfile + " does not exist!")
        #    exit
        # check for dbtable
        conn = sqlite3.connect(dbfile)
        curs = conn.cursor()
        list_tables_sql = '''SELECT name FROM sqlite_master
            WHERE type='table' and name='%s'
            ORDER BY name;''' % dbtable
        curs.execute(list_tables_sql)
        list_tables = curs.fetchall()
        if list_tables == '':
            exit
        # vascular plants
        elif list_tables[0][0] == 'dao_pnamelist' or list_tables[0][0] == 'dao_pnamelist_pg' \
                or list_tables[0][0] == 'dao_jp_ylist':
            species_type = 1
        # birds
        elif list_tables[0][0] == 'dao_bnamelist':
            species_type = 2
        else:
            species_type = 1

        #### INPUT FILES
        self.importTable(dbfile, table_name = 'sample', import_file=inputfile)
        # curs.execute('DROP TABLE IF EXISTS sample;')
        # conn.commit()
        # sample_create = '''
        # CREATE TABLE sample (
        #   cname varchar
        # );
        # '''
        # curs.execute(sample_create)
        # with open(inputfile, newline='', encoding='utf-8') as f:
        #     reader = csv.reader(f, delimiter='|')
        #     for row in reader:
        #         # substitute 台 to 臺
        #         zhname = re.sub('台([灣|北|中|西|南|東])',r'臺\1', row[0])
        #         insert_db = '''
        #         INSERT INTO sample (cname) VALUES ("%s");
        #         ''' % zhname
        #         curs.execute(insert_db)
        #         conn.commit()
        # f.close()

        ##### EXPORT DwC ##### 
        query_all = '''
            SELECT 
                distinct n.id, n.family, n.family_cname,
                n.fullname,n.cname,n.endemic,n.iucn_category,n.source
            FROM sample s LEFT OUTER JOIN
            (SELECT distinct plant_type, id, family,
                family_cname,fullname, cname,
                endemic, iucn_category, source
            FROM %s ORDER BY plant_type, family) as n ON s.local_name = n.cname 
            ORDER BY n.plant_type, n.family, n.fullname;
        ''' % ( dbtable )
        curs.execute(query_all)
        dwc_list = curs.fetchall()
        with codecs.open(ofile_prefix + '.csv', 'w+', 'utf-8') as f:
            writer = csv.writer(f, delimiter='\t', doublequote=True, quoting=1, dialect='excel')
            writer.writerow(['taxonID','family', 'familyVernacularName', 'scientificName', \
                                        'vernacularName', 'isEndemic', 'iucnCategory', 'source'])
            for i in range(0, len(dwc_list)):
                writer.writerow(dwc_list[i])

        ##### EXPORT Markdown #####
        with codecs.open(ofile_prefix +'.md', 'w+', 'utf-8') as f:
            ##### Generate HEADER #####
            if species_type == 1:
                f.write(u'# 維管束植物名錄')
                sp_note = u'"#" 代表特有種，"*" 代表歸化種，"†" 代表栽培種。'
                sp_conserv = u'''中名後面括號內的縮寫代表依照「臺灣維管束植物紅皮書初評名錄」中依照 IUCN 瀕危物種所評估等級， \
EX: 滅絕、EW: 野外滅絕、RE: 區域性滅絕、CR: 嚴重瀕臨滅絕、 \
EN: 瀕臨滅絕、VU: 易受害、NT: 接近威脅、DD: 資料不足。若未註記者代表安全(Least concern)'''
            elif species_type == 2:
                f.write(u'# 鳥類名錄')
                sp_note = u'"#" 代表特有種，"##" 代表特有亞種'
                sp_conserv = u'''中名後面括號內代表行政院農業委員會依照野生動物保護法所公布之保育等級。 \
I：表示瀕臨絕種野生動物、II：表示珍貴稀有野生動物、III：表示其他應予保育之野生動物'''
            else:
                f.write(u'# 物種名錄')
            f.write('\n')
            ##### End of HEADER #####
            count_family = '''
            SELECT count(*) from (SELECT distinct family from sample s left outer join %s n
                    on s.local_name=n.cname) as f;
            ''' % dbtable
            count_species = '''
            SELECT count(*) from (SELECT distinct n.cname from sample s left outer join %s n
                    on s.local_name=n.cname) as f;
            ''' % dbtable
            not_exist_sp = '''
            SELECT distinct s.local_name from sample s left outer join %s n
                    on s.local_name=n.cname where n.cname is null;
            ''' % dbtable
            curs.execute(count_family)
            family_no = curs.fetchall()[0][0]
            curs.execute(count_species)
            species_no = curs.fetchall()[0][0]
            curs.execute(not_exist_sp)
            no_sp = curs.fetchall()
            nsp = []
            for i in no_sp:
                nsp.append(i[0])
            nsp = ', '.join(nsp)
            if len(nsp) > 0:
                f.write('\n')
                f.write(u'<font color="red">輸入名錄中，下列物種不存在於物種資料庫中：{} ，請再次確認物種中名是否和資料庫中相同</font>\n'.format(nsp))
            f.write('\n')
            f.write(u'本名錄中共有 {} 科、{} 種，科名後括弧內為該科之物種總數。'.format(family_no, species_no))
            f.write(sp_note)
            f.write(sp_conserv)
            f.write('\n')
            ####### End of HEADER

            ####### namelist BODY
            pt_plant_type_sql = '''
                SELECT p.plant_type,p.pt_name
                FROM dao_plant_type p,
                    (SELECT distinct plant_type from sample s left outer join %s n
                    on s.local_name=n.cname order by plant_type) as t
                WHERE p.plant_type = t.plant_type;
            ''' % dbtable
            curs.execute(pt_plant_type_sql)
            pt_plant_type = curs.fetchall()
            n = 1
            m = 1

            for i in range(0,len(pt_plant_type)):
                f.write('\n')
                f.write('\n###'+pt_plant_type[i][1]+'\n\n')
                taxa_family_sql = '''
                select distinct family,family_cname from sample s left outer join %s n
                on s.local_name=n.cname where n.plant_type=%i
                order by plant_type,family;
                ''' % (dbtable, pt_plant_type[i][0])
                curs.execute(taxa_family_sql)
                taxa_family = curs.fetchall()
                for j in range(0,len(taxa_family)):
                    sp_number_in_fam = '''
                    select count(*) from
                        (select distinct fullname, name, n.cname from sample s left outer join %s n
                        on s.local_name=n.cname where n.plant_type=%i and family='%s'
                        order by plant_type,family,fullname) as a;
                    ''' % (dbtable, pt_plant_type[i][0], taxa_family[j][0])
                    curs.execute(sp_number_in_fam)
                    fam_spno = curs.fetchall()[0][0]
                    fam = str(m) + '. **' + taxa_family[j][0]
                    fam_zh = taxa_family[j][1] + '**'
                    f.write('\n')
                    f.write(fam + ' ' + fam_zh + ' (%i)\n' % fam_spno)
                    taxa_family_sp = '''
                        select distinct fullname,n.cname,n.endemic,n.source,n.iucn_category,n.name from sample s left outer join %s n
                        on s.local_name=n.cname where n.plant_type=%i and family='%s'
                        order by plant_type,family,fullname;
                    ''' % (dbtable, pt_plant_type[i][0], taxa_family[j][0])
                    curs.execute(taxa_family_sp)
                    taxa_family_sp = curs.fetchall()
                    m = m + 1
                    # output species within a family
                    for k in range(0,len(taxa_family_sp)):
                        # check the endmic species
                        if taxa_family_sp[k][2] == 1:
                            ENDEMIC = "#"
                        else:
                            ENDEMIC = ''
                        # check the source
                        if taxa_family_sp[k][3] == u'栽培':
                            SRC = '†'
                        elif taxa_family_sp[k][3] == u'歸化':
                            SRC = '*'
                        else:
                            SRC = ''
                        # IUCN category
                        if len(taxa_family_sp[k][4]) == 2:
                            IUCNCAT = ' (%s)' % taxa_family_sp[k][4]
                        else:
                            IUCNCAT = ''
                        spinfo = ' ' + ENDEMIC + SRC + IUCNCAT

                        if spinfo is not None:
                            f.write('    ' + str(n) + '. ' + self.fmtname(taxa_family_sp[k][0], split=False) \
                                        + ' ' + taxa_family_sp[k][1] + spinfo + '\n')
                        else:
                            f.write('    ' + str(n) + '. ' + self.fmtname(taxa_family_sp[k][0], split=False) \
                                    + ' ' + taxa_family_sp[k][1] +'\n')
                        n = n + 1
            f.close()

            try:
                if oformat == 'xlsx':
                    dwc_list.insert(0, ['taxonID','family', 'familyVernacularName', 'Fullname with Authors', \
                                        'vernacularName', 'isEndemic', 'iucnCategory', 'source'])
                    self.listToXls(dwc_list, 3, ofile_prefix)
                else:
                    self.pandocConvert(oformat, ofile_prefix)
            except BaseException as e:
                print(str(e))
            # cleanup everything
            curs.execute('DROP TABLE IF EXISTS sample;')
            conn.commit()
            conn.close()

if __name__ == "__main__":
    import doctest
    doctest.testmod()
